import pandas as pd
from openpyxl import load_workbook

def read_excel_in_scratch(file_path, sheet_name="Q1-Q2"):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []
    columns = [cell.value for cell in sheet[1]]  # Получаем заголовки из первой строки
    for row in sheet.iter_rows(min_row=2):  # Пропускаем заголовок, возвращаем объекты ячеек
        l = []
        for i in range(len(row)):
            cell = row[i]
            if cell.hyperlink:
                l.append(f"[{cell.value}]({cell.hyperlink.target})")
            else:
                l.append(cell.value)
        # journal_cell = row[0]
        # q = row[1].value
        # apc_cell = row[2]
        # comments = row[3].value

        # journal_link = journal_cell.hyperlink.target if journal_cell.hyperlink else None
        # apc_link = apc_cell.hyperlink.target if apc_cell.hyperlink else None

        # journal_md = f"[{journal_cell.value}]({journal_link})" if journal_link else journal_cell.value
        # apc_md = f"[{apc_cell.value}]({apc_link})" if apc_link else apc_cell.value

        # data.append([journal_md, q, apc_md, comments])
        if any(value is not None for value in l):
            data.append(l)
    df = pd.DataFrame(data, columns=columns)
    df = df.dropna(how='all')
    return df

def create_readme():
    # df = pd.read_excel('./source/Journals.xlsx', sheet_name="Q1-Q2", 
    #                    na_values="", dtype=str)
    df = read_excel_in_scratch('./source/Journals.xlsx', sheet_name="Q1-Q2")
    df['Journal'] = "[" + df['Journal'] + "](" + df['Link'] + ")"
    markdown_table = df[["Journal", "Q", "APC", "Comments + Turnaround time"]].to_markdown(index=False)

    with open('./source/Canvas.md', 'r') as file:
        content = file.readlines()

    start_index = content.index("## Q1-Q2 Journals\n") + 1
    content.insert(start_index, "\n### Main Journal Table\n" + markdown_table + "\n\n")

    with open('./README.md', 'w') as file:
        file.writelines(content)

if __name__ == "__main__":
    create_readme()